from functools import wraps
from base64 import b64encode
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from tempfile import NamedTemporaryFile
from wtforms import Form, StringField, DateField, validators, RadioField
from flask import Flask, render_template, redirect, request, url_for, flash, jsonify, send_from_directory, send_file
from flask_login import LoginManager, login_user, current_user
from peewee import DoesNotExist

from models import User, Activity
from config import ADMIN, ANY, MONTH_NUMBERS

app = Flask(__name__)
app.config.from_pyfile('config.py')
app.secret_key = 'secret'

lm = LoginManager()
lm.init_app(app)


class ActivityForm(Form):
    name = StringField('Назва', [validators.DataRequired()])
    description = StringField('Опис', [validators.Optional()])

    activity_type = RadioField('Тип', [validators.DataRequired()],
                               choices=[('organized', 'Організував, провів, зробив'),
                                        ('collaborated', 'Долучився'),
                                        ('visited', 'Відвідав')])

    date = DateField('Дата', [validators.DataRequired()])


class UserForm(Form):
    name = StringField('Ім`я', [validators.DataRequired()])


def login_required(role=ANY):
    def wrapper(fn):
        @wraps(fn)
        def decorated_view(*args, **kwargs):
            if not current_user.is_authenticated:
                return lm.unauthorized()

            if role == ANY:
                return fn(*args, **kwargs)

            elif current_user.role < role:
                return lm.unauthorized()

            else:
                return fn(*args, **kwargs)

        return decorated_view
    return wrapper


@lm.user_loader
def load_user(user_id):
    try:
        return User.get(code=user_id)
    except DoesNotExist:
        return


@app.route("/")
def index():
    if not current_user.is_authenticated:
        return redirect(url_for('auth'))

    return render_template('main.html')


@app.route('/auth')
def auth():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    return render_template('auth.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    def bad_login_redirect():
        flash('Введений код невірний!')
        return redirect(url_for('index'))

    if not request.method == 'POST':
        return redirect(url_for('index'))

    code = request.form['code']

    try:
        user = User.get(code=code)

    except DoesNotExist:
        return bad_login_redirect()

    login_user(user)
    user.authenticated = True
    user.save()
    return redirect(url_for('index'))


@app.route('/add_activity', methods=['POST'])
@login_required()
def add_activity():
    form = ActivityForm(request.form)
    if form.validate():
        name, description, activity_type, date = form.name.data, form.description.data, \
                                                 form.activity_type.data, form.date.data

        Activity.create(user=current_user.id, name=name, date=date,
                        description=description, activity_type=activity_type)

        flash('Звіт надіслано!')
        return redirect(url_for('index'))

    else:
        flash('Невірно заповнена форма')
        return redirect(url_for('index'))


@app.route('/admin')
@login_required(ADMIN)
def admin():
    return render_template('admin.html', users=User.select().order_by(User.name))


@app.route('/add_user', methods=['POST'])
@login_required(ADMIN)
def add_user():
    form = UserForm(request.form)
    if form.validate():
        name = form.name.data

        User.create(name=name, code=User.generate_code())

        flash('Користувач доданий!')
        return redirect(url_for('admin'))

    else:
        flash('Невірно заповнена форма')
        return redirect(url_for('admin'))


@app.route('/delete_user/<int:user_id>')
@login_required()
def delete_user(user_id):
    try:
        user = User.get(id=user_id)
        user.delete_instance()
        flash('Члена видалено!')
        return jsonify({'success': True, 'message': 'Члена видалено!'})

    except User.DoesNotExist:
        return jsonify({'success': False, 'message': 'Не знайдено'})


@app.route('/get_qr/<int:user_id>')
@login_required(ADMIN)
def get_qr(user_id):
    try:
        user = User.get(id=user_id)
        qrcode = user.get_qrcode()

        return jsonify({'success': True, 'qrcode': b64encode(qrcode).decode('ascii'), 'code': user.code, 'name': user.name})

    except User.DoesNotExist:
        return jsonify({'success': False, 'msg': 'Такого користувача не існує!'})


@app.route('/save_roles', methods=['POST'])
@login_required(ADMIN)
def save_roles():
    json = request.get_json()
    if json:
        for user_role in json:
            pk, role = int(user_role['_pk']), int(user_role['role'])
            if 0 <= role < 2:
                user = User.get(id=pk)
                user.role = role
                user.save()

        return jsonify({'success': True, 'msg': 'Привілегії змінено!'})


def generate_xlsx(month):
    def as_text(value):
        if value is None:
            return ""
        return str(value)

    wb = Workbook()
    ws = wb.active

    alignment = Alignment(horizontal='general',
                          vertical='top',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

    ws.append(['№', 'ПІБ', 'Організував, провів, зробив', 'Долучився', 'Відвідав',
               'Чергове засідання', 'Позачергові засідання'])

    users = User.select().order_by(User.name)
    activities = Activity.select().where(Activity.date.month == month)

    for i, user in enumerate(users):
        organized = activities.join(User).where((Activity.activity_type == 'organized') &
                                                (User.name == user.name))

        collaborated = activities.join(User).where((Activity.activity_type == 'collaborated') &
                                                   (User.name == user.name))

        visited = activities.join(User).where((Activity.activity_type == 'visited') &
                                              (User.name == user.name))

        organized_str = ""
        for j, act in enumerate(organized):
            organized_str += f'{j+1}. {act.name} [{act.date.strftime("%d.%m.%Y")}]'

            if act.description:
                organized_str += f'\nОпис: {act.description}\n\n'
            else:
                organized_str += '\n'

        collaborated_str = ""
        for j, act in enumerate(collaborated):
            collaborated_str += f'{j+1}. {act.name} [{act.date.strftime("%d.%m.%Y")}]'

            if act.description:
                collaborated_str += f'\nОпис: {act.description}\n\n'
            else:
                collaborated_str += '\n'

        visited_str = ""
        for j, act in enumerate(visited):
            visited_str += f'{j+1}. {act.name} [{act.date.strftime("%d.%m.%Y")}]'

            if act.description:
                visited_str += f'\nОпис: {act.description}\n\n'
            else:
                visited_str += '\n'

        ws.append([str(i+1), user.name, organized_str.strip(),
                   collaborated_str.strip(), visited_str.strip()])

    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:

            max_length_of_line_in_cell = max(len(line)
                                             for line in as_text(cell.value)
                                             .strip()
                                             .split('\n'))

            max_length = max(max_length_of_line_in_cell, max_length)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2

    for row_cells in ws.rows:
        height = max(as_text(cell.value).count('\n') for cell in row_cells)
        ws.row_dimensions[row_cells[0].row].height = 11*(height+2)

    max_length = 0
    max_height = 0

    for col in ws.columns:
        for cell in col:
            val = as_text(cell.value)
            cell.alignment = alignment

            # TODO: Add normal autosize
            max_length = max(len(val), max_length)
            max_height = max(val.count('\n'), max_height)

    tf = NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tf.name)

    return tf.name.replace('/tmp/', '')


@app.route('/get_xlsx/<string:month_name>')
@login_required()
def get_xlsx(month_name):
    if month_name in MONTH_NUMBERS.keys():
        filename = generate_xlsx(MONTH_NUMBERS[month_name])
        return send_from_directory('/tmp/', filename, as_attachment=True)
    else:
        return jsonify({'success': False, 'message': 'Помилка'})


@app.route('/get_activities/<string:month_name>')
@login_required()
def get_stats(month_name):
    activities = current_user.activities.select()\
        .where(Activity.date.month == MONTH_NUMBERS[month_name])\
        .order_by(Activity.date.desc())

    return render_template('dynamic/activities.html', activities=activities, verbose_type={
        'organized': 'Організував',
        'collaborated': 'Долучився',
        'visited': 'Відвідав'
    })


@app.route('/delete_activity/<int:activity_id>')
@login_required()
def delete_activity(activity_id):
    try:
        activity = Activity.get(id=activity_id)
    except Activity.DoesNotExist:
        return jsonify({'success': False, 'message': 'Не знайдено'})

    if activity.user.id != current_user.id:
        return jsonify({'success': False, 'message': 'Помилка доступу'})
    else:
        activity.delete_instance()
        return jsonify({'success': True, 'message': 'Видалено'})


@app.route('/stats')
@login_required()
def stats():
    return render_template('stats.html')


@app.route('/uploads/<string:filename>')
def download_file(filename):
    return send_from_directory('/tmp/', filename, as_attachment=True)
